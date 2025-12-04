"""
Universal file format processor - handles ANY file type
"""

import os
import zipfile
import tarfile
import rarfile
import py7zr
import pickle
import json
import yaml
import xml.etree.ElementTree as ET
import csv
import pdfplumber
import docx
import openpyxl
import pytesseract
from PIL import Image
import speech_recognition as sr
import pydub
import olefile
import lief  # Binary analysis
import capstone  # Disassembler
import keystone  # Assembler
import networkx as nx
import torch
import tensorflow as tf
import joblib
import h5py
import numpy as np
from pathlib import Path
import subprocess
import magic
import filetype

class OmniversalAdaptor:
    """Processes ANY file type automatically"""
    
    def __init__(self, watch_folder="training_data"):
        self.watch_folder = Path(watch_folder)
        self.watch_folder.mkdir(exist_ok=True)
        
        # Supported file type processors
        self.processors = {
            'archive': self.process_archive,
            'executable': self.process_executable,
            'document': self.process_document,
            'image': self.process_image,
            'audio': self.process_audio,
            'video': self.process_video,
            'code': self.process_code,
            'data': self.process_data,
            'model': self.process_model,
            'database': self.process_database,
            'config': self.process_config,
            'unknown': self.process_unknown
        }
        
        # Start watcher
        self.start_file_watcher()
    
    def start_file_watcher(self):
        """Watch for new files every 60 seconds"""
        import threading
        import time
        
        def watcher():
            processed_files = set()
            
            while True:
                try:
                    # Get all files
                    all_files = list(self.watch_folder.rglob("*"))
                    
                    for file_path in all_files:
                        if file_path.is_file() and file_path not in processed_files:
                            print(f"[+] New file detected: {file_path}")
                            
                            # Process file
                            self.process_file(file_path)
                            
                            # Add to processed set
                            processed_files.add(file_path)
                            
                            # If it's an archive, also process contents
                            if self.is_archive(file_path):
                                self.extract_and_process(file_path)
                    
                    time.sleep(60)  # Check every 60 seconds
                    
                except Exception as e:
                    print(f"[-] Watcher error: {e}")
                    time.sleep(10)
        
        thread = threading.Thread(target=watcher, daemon=True)
        thread.start()
    
    def process_file(self, file_path):
        """Process any file type"""
        try:
            # Determine file type
            file_type = self.detect_file_type(file_path)
            print(f"[+] Processing {file_type}: {file_path.name}")
            
            # Get appropriate processor
            processor = self.processors.get(file_type, self.processors['unknown'])
            
            # Process file
            result = processor(file_path)
            
            # Integrate learned capabilities
            self.integrate_capabilities(result, file_path)
            
            return result
            
        except Exception as e:
            print(f"[-] Error processing {file_path}: {e}")
            return None
    
    def detect_file_type(self, file_path):
        """Detect file type using multiple methods"""
        # Use file magic
        mime = magic.Magic(mime=True)
        mime_type = mime.from_file(str(file_path))
        
        # Check extensions
        ext = file_path.suffix.lower()
        
        # Archive types
        archive_exts = {'.zip', '.tar', '.gz', '.bz2', '.xz', '.7z', '.rar', '.tgz'}
        if ext in archive_exts or 'archive' in mime_type:
            return 'archive'
        
        # Executable types
        exec_exts = {'.exe', '.dll', '.so', '.dylib', '.bin', '.elf', '.py', '.sh', '.bat', '.jar'}
        if ext in exec_exts or 'executable' in mime_type or 'application/x-' in mime_type:
            return 'executable'
        
        # Document types
        doc_exts = {'.pdf', '.doc', '.docx', '.txt', '.rtf', '.odt', '.md'}
        if ext in doc_exts or 'text/' in mime_type or 'document' in mime_type:
            return 'document'
        
        # Image types
        image_exts = {'.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.svg'}
        if ext in image_exts or 'image/' in mime_type:
            return 'image'
        
        # Audio types
        audio_exts = {'.mp3', '.wav', '.flac', '.aac', '.ogg', '.m4a'}
        if ext in audio_exts or 'audio/' in mime_type:
            return 'audio'
        
        # Video types
        video_exts = {'.mp4', '.avi', '.mov', '.wmv', '.flv', '.mkv'}
        if ext in video_exts or 'video/' in mime_type:
            return 'video'
        
        # Code types
        code_exts = {'.py', '.js', '.java', '.cpp', '.c', '.h', '.cs', '.php', '.rb', '.go', '.rs'}
        if ext in code_exts:
            return 'code'
        
        # Data types
        data_exts = {'.json', '.xml', '.yaml', '.yml', '.csv', '.tsv', '.h5', '.pkl', '.joblib'}
        if ext in data_exts:
            return 'data'
        
        # Model types
        model_exts = {'.pt', '.pth', '.h5', '.keras', '.onnx', '.tflite'}
        if ext in model_exts:
            return 'model'
        
        # Database types
        db_exts = {'.db', '.sqlite', '.sqlite3', '.mdb', '.accdb'}
        if ext in db_exts:
            return 'database'
        
        # Config types
        config_exts = {'.ini', '.cfg', '.conf', '.config', '.properties'}
        if ext in config_exts:
            return 'config'
        
        return 'unknown'
    
    def process_archive(self, file_path):
        """Process archive files"""
        extracted_files = []
        
        try:
            # Extract based on archive type
            if file_path.suffix == '.zip':
                with zipfile.ZipFile(file_path, 'r') as zip_ref:
                    extract_path = self.watch_folder / file_path.stem
                    zip_ref.extractall(extract_path)
                    extracted_files = list(extract_path.rglob("*"))
                    
            elif file_path.suffix in ['.tar', '.tgz', '.gz', '.bz2', '.xz']:
                with tarfile.open(file_path, 'r:*') as tar_ref:
                    extract_path = self.watch_folder / file_path.stem
                    tar_ref.extractall(extract_path)
                    extracted_files = list(extract_path.rglob("*"))
                    
            elif file_path.suffix == '.7z':
                with py7zr.SevenZipFile(file_path, 'r') as sevenz_ref:
                    extract_path = self.watch_folder / file_path.stem
                    sevenz_ref.extractall(extract_path)
                    extracted_files = list(extract_path.rglob("*"))
                    
            elif file_path.suffix == '.rar':
                with rarfile.RarFile(file_path, 'r') as rar_ref:
                    extract_path = self.watch_folder / file_path.stem
                    rar_ref.extractall(extract_path)
                    extracted_files = list(extract_path.rglob("*"))
            
            # Process extracted files
            for extracted_file in extracted_files:
                if extracted_file.is_file():
                    self.process_file(extracted_file)
            
            return extracted_files
            
        except Exception as e:
            print(f"[-] Error extracting {file_path}: {e}")
            return []
    
    def process_executable(self, file_path):
        """Analyze executable files"""
        analysis = {
            'file': str(file_path),
            'type': 'executable',
            'analysis': {}
        }
        
        try:
            # Binary analysis with LIEF
            binary = lief.parse(str(file_path))
            if binary:
                analysis['analysis']['binary'] = {
                    'format': str(binary.format),
                    'entrypoint': hex(binary.entrypoint),
                    'sections': [{
                        'name': section.name,
                        'size': section.size,
                        'virtual_address': hex(section.virtual_address)
                    } for section in binary.sections],
                    'imports': [str(imp) for imp in binary.imports],
                    'exports': [str(exp) for exp in binary.exports]
                }
            
            # Disassembly with Capstone
            with open(file_path, 'rb') as f:
                code = f.read()
                
            md = capstone.Cs(capstone.CS_ARCH_X86, capstone.CS_MODE_64)
            instructions = []
            
            for instruction in md.disasm(code, 0x1000):
                instructions.append(f"{instruction.mnemonic} {instruction.op_str}")
                
                if len(instructions) > 100:  # Limit
                    break
            
            analysis['analysis']['disassembly'] = instructions[:50]
            
            # Strings extraction
            import re
            strings = re.findall(b'[\\x20-\\x7E]{4,}', code)
            analysis['analysis']['strings'] = [s.decode('utf-8', errors='ignore') for s in strings[:20]]
            
            # If it's Python, analyze source
            if file_path.suffix == '.py':
                with open(file_path, 'r') as f:
                    source = f.read()
                analysis['analysis']['source'] = source[:1000]  # First 1000 chars
            
            return analysis
            
        except Exception as e:
            print(f"[-] Error analyzing executable {file_path}: {e}")
            return analysis
    
    def process_document(self, file_path):
        """Extract text from documents"""
        text = ""
        
        try:
            if file_path.suffix == '.pdf':
                with pdfplumber.open(file_path) as pdf:
                    for page in pdf.pages:
                        text += page.extract_text() + "\\n"
                        
            elif file_path.suffix in ['.docx', '.doc']:
                import docx
                doc = docx.Document(file_path)
                text = "\\n".join([para.text for para in doc.paragraphs])
                
            elif file_path.suffix in ['.xlsx', '.xls']:
                wb = openpyxl.load_workbook(file_path)
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    for row in ws.iter_rows(values_only=True):
                        text += " ".join([str(cell) for cell in row if cell]) + "\\n"
                        
            elif file_path.suffix == '.txt':
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    text = f.read()
            
            # Analyze text for patterns
            analysis = self.analyze_text_patterns(text)
            
            return {
                'file': str(file_path),
                'type': 'document',
                'text': text[:5000],  # Limit size
                'analysis': analysis
            }
            
        except Exception as e:
            print(f"[-] Error processing document {file_path}: {e}")
            return {'file': str(file_path), 'error': str(e)}
    
    def process_image(self, file_path):
        """Extract text and analyze images"""
        try:
            # OCR with Tesseract
            image = Image.open(file_path)
            text = pytesseract.image_to_string(image)
            
            # Analyze image properties
            width, height = image.size
            mode = image.mode
            
            return {
                'file': str(file_path),
                'type': 'image',
                'dimensions': f"{width}x{height}",
                'mode': mode,
                'ocr_text': text
            }
            
        except Exception as e:
            print(f"[-] Error processing image {file_path}: {e}")
            return {'file': str(file_path), 'error': str(e)}
    
    def process_code(self, file_path):
        """Analyze code files"""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                code = f.read()
            
            # Extract imports, functions, classes
            imports = []
            functions = []
            classes = []
            
            lines = code.split('\\n')
            for line in lines:
                line = line.strip()
                
                # Find imports
                if line.startswith('import ') or line.startswith('from '):
                    imports.append(line)
                
                # Find function definitions
                elif line.startswith('def '):
                    functions.append(line)
                
                # Find class definitions
                elif line.startswith('class '):
                    classes.append(line)
            
            # Security analysis
            security_issues = self.analyze_code_security(code)
            
            return {
                'file': str(file_path),
                'type': 'code',
                'language': file_path.suffix[1:],
                'imports': imports,
                'functions': functions[:10],  # Limit
                'classes': classes[:10],  # Limit
                'security_issues': security_issues,
                'size': len(code)
            }
            
        except Exception as e:
            print(f"[-] Error processing code {file_path}: {e}")
            return {'file': str(file_path), 'error': str(e)}
    
    def process_model(self, file_path):
        """Load and analyze ML models"""
        try:
            model_info = {
                'file': str(file_path),
                'type': 'model',
                'framework': None,
                'architecture': None,
                'parameters': None
            }
            
            # PyTorch models
            if file_path.suffix in ['.pt', '.pth']:
                model = torch.load(file_path, map_location='cpu')
                model_info['framework'] = 'pytorch'
                
                if isinstance(model, dict):
                    model_info['architecture'] = list(model.keys())
                    # Estimate parameters
                    total_params = 0
                    for key, value in model.items():
                        if hasattr(value, 'numel'):
                            total_params += value.numel()
                    model_info['parameters'] = total_params
            
            # TensorFlow/Keras models
            elif file_path.suffix in ['.h5', '.keras']:
                import tensorflow as tf
                model = tf.keras.models.load_model(file_path)
                model_info['framework'] = 'tensorflow'
                model_info['architecture'] = [layer.__class__.__name__ for layer in model.layers]
                model_info['parameters'] = model.count_params()
            
            # ONNX models
            elif file_path.suffix == '.onnx':
                import onnx
                model = onnx.load(str(file_path))
                model_info['framework'] = 'onnx'
                model_info['architecture'] = [node.op_type for node in model.graph.node]
            
            return model_info
            
        except Exception as e:
            print(f"[-] Error processing model {file_path}: {e}")
            return {'file': str(file_path), 'error': str(e)}
    
    def process_unknown(self, file_path):
        """Attempt to process unknown file types"""
        try:
            # Try to read as binary
            with open(file_path, 'rb') as f:
                data = f.read(1024)  # First 1KB
            
            # Analyze binary patterns
            hex_dump = data.hex()[:200]
            
            # Try to detect if it's encrypted/compressed
            entropy = self.calculate_entropy(data)
            
            return {
                'file': str(file_path),
                'type': 'unknown',
                'size': os.path.getsize(file_path),
                'entropy': entropy,
                'hex_sample': hex_dump,
                'likely_type': self.guess_file_type(data)
            }
            
        except Exception as e:
            return {'file': str(file_path), 'error': str(e)}
    
    def integrate_capabilities(self, result, file_path):
        """Integrate learned capabilities into AI"""
        if not result:
            return
        
        # Save analysis
        analysis_dir = Path("capabilities") / "learned"
        analysis_dir.mkdir(parents=True, exist_ok=True)
        
        analysis_file = analysis_dir / f"{file_path.stem}_analysis.json"
        
        with open(analysis_file, 'w') as f:
            json.dump(result, f, indent=2)
        
        # Check if we should integrate code
        if 'code' in str(result.get('type', '')):
            self.integrate_code_capabilities(result, file_path)
        
        # Check if we should integrate model
        if 'model' in str(result.get('type', '')):
            self.integrate_model_capabilities(result, file_path)
    
    def integrate_code_capabilities(self, result, file_path):
        """Integrate code capabilities"""
        # Extract useful functions/classes
        if 'functions' in result:
            # Create wrapper for useful functions
            code_dir = Path("capabilities") / "code_library"
            code_dir.mkdir(exist_ok=True)
            
            # Copy the file if it looks useful
            if self.is_useful_code(result):
                shutil.copy(file_path, code_dir / file_path.name)
                
                # Try to import and test
                self.test_and_integrate_module(file_path)
    
    def integrate_model_capabilities(self, result, file_path):
        """Integrate ML model capabilities"""
        model_dir = Path("capabilities") / "models"
        model_dir.mkdir(exist_ok=True)
        
        # Copy model
        shutil.copy(file_path, model_dir / file_path.name)
        
        # Create wrapper
        wrapper = f'''
# Auto-generated wrapper for {file_path.name}
import torch
import numpy as np

class ModelWrapper:
    def __init__(self, model_path="{model_dir / file_path.name}"):
        self.model = torch.load(model_path, map_location='cpu')
        self.model.eval()
    
    def predict(self, input_data):
        with torch.no_grad():
            return self.model(input_data)
    
    def extract_features(self, input_data):
        # Extract intermediate features
        features = []
        def hook(module, input, output):
            features.append(output)
        
        # Register hooks
        handles = []
        for name, layer in self.model.named_children():
            handles.append(layer.register_forward_hook(hook))
        
        # Forward pass
        self.model(input_data)
        
        # Remove hooks
        for handle in handles:
            handle.remove()
        
        return features
'''
        
        wrapper_path = model_dir / f"wrapper_{file_path.stem}.py"
        wrapper_path.write_text(wrapper)
    
    def is_useful_code(self, analysis):
        """Determine if code is useful for capabilities"""
        useful_keywords = [
            'exploit', 'vulnerability', 'scan', 'crawl', 'spider',
            'bruteforce', 'crack', 'hash', 'encrypt', 'decrypt',
            'network', 'socket', 'request', 'response', 'api',
            'database', 'sql', 'nosql', 'injection', 'xss',
            'rce', 'lfi', 'rfi', 'ssrf', 'csrf', 'xxe',
            'deserialization', 'buffer', 'overflow', 'shell',
            'reverse', 'payload', 'malware', 'trojan', 'virus',
            'worm', 'ransomware', 'keylogger', 'rootkit',
            'persistence', 'privilege', 'escalation', 'bypass',
            'evasion', 'stealth', 'obfuscation', 'packer',
            'crypter', 'fud', 'undetectable', 'antivirus',
            'firewall', 'ids', 'ips', 'waf', 'bypass'
        ]
        
        code_text = str(analysis).lower()
        
        for keyword in useful_keywords:
            if keyword in code_text:
                return True
        
        return False
    
    def test_and_integrate_module(self, file_path):
        """Test and integrate a Python module"""
        try:
            # Create temporary test
            test_script = f'''
import sys
sys.path.insert(0, "{file_path.parent}")

try:
    import {file_path.stem} as module
    print(f"[+] Successfully imported {file_path.stem}")
    
    # Test basic functionality
    if hasattr(module, 'main'):
        module.main()
    elif hasattr(module, 'run'):
        module.run()
    
except Exception as e:
    print(f"[-] Error testing {file_path.stem}: {{e}}")
'''
            
            # Run test
            result = subprocess.run([sys.executable, '-c', test_script], 
                                  capture_output=True, text=True)
            
            if "Successfully imported" in result.stdout:
                print(f"[+] Integrated module: {file_path.stem}")
                
                # Add to capabilities registry
                registry_path = Path("capabilities") / "registry.json"
                if registry_path.exists():
                    with open(registry_path, 'r') as f:
                        registry = json.load(f)
                else:
                    registry = {"modules": []}
                
                registry["modules"].append({
                    "name": file_path.stem,
                    "path": str(file_path),
                    "integrated": datetime.now().isoformat()
                })
                
                with open(registry_path, 'w') as f:
                    json.dump(registry, f, indent=2)
        
        except Exception as e:
            print(f"[-] Error testing module {file_path}: {e}")
    
    def analyze_text_patterns(self, text):
        """Analyze text for useful patterns"""
        patterns = {
            'urls': [],
            'emails': [],
            'ip_addresses': [],
            'phone_numbers': [],
            'crypto_addresses': [],
            'api_keys': [],
            'passwords': [],
            'usernames': []
        }
        
        import re
        
        # URLs
        patterns['urls'] = re.findall(r'https?://[^\s<>"\']+|www\.[^\s<>"\']+', text)
        
        # Emails
        patterns['emails'] = re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', text)
        
        # IP addresses
        patterns['ip_addresses'] = re.findall(r'\b\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}\b', text)
        
        # Phone numbers
        patterns['phone_numbers'] = re.findall(r'\b\d{3}[-.]?\d{3}[-.]?\d{4}\b', text)
        
        # Crypto addresses (basic patterns)
        patterns['crypto_addresses'] = re.findall(r'\b(0x)?[0-9a-fA-F]{40}\b', text)
        
        # API keys (common patterns)
        patterns['api_keys'] = re.findall(r'\b[A-Za-z0-9_]{32,}\b', text)
        
        return patterns
    
    def analyze_code_security(self, code):
        """Analyze code for security issues"""
        issues = []
        
        # Dangerous patterns
        dangerous_patterns = [
            (r'eval\s*\(', 'eval() function - code injection risk'),
            (r'exec\s*\(', 'exec() function - code injection risk'),
            (r'__import__\s*\(', 'dynamic import - potential malware'),
            (r'os\.system\s*\(', 'os.system() - command injection risk'),
            (r'subprocess\.Popen\s*\(', 'subprocess.Popen - command injection risk'),
            (r'pickle\.loads\s*\(', 'pickle.loads() - deserialization attack'),
            (r'yaml\.load\s*\(', 'yaml.load() - deserialization attack'),
            (r'input\s*\(', 'input() - potential injection if not validated'),
            (r'str\.format\s*\(.*\{.*\}.*\)', 'Potential format string vulnerability'),
            (r'sql\s*\+', 'String concatenation in SQL - injection risk'),
            (r'http://|https://.*\{.*\}', 'Potential SSRF if user-controlled'),
            (r'File\.open\s*\(.*\{.*\}', 'Potential path traversal'),
            (r'requests\.get\s*\(.*\{.*\}', 'Potential SSRF'),
            (r'rm\s+-rf|del\s+/f|shutil\.rmtree', 'Dangerous delete operations'),
            (r'base64\.b64decode', 'Base64 decoding - could be obfuscated payload')
        ]
        
        for pattern, description in dangerous_patterns:
            if re.search(pattern, code, re.IGNORECASE):
                issues.append(description)
        
        return issues
    
    def calculate_entropy(self, data):
        """Calculate entropy of data"""
        if not data:
            return 0
        
        entropy = 0
        for x in range(256):
            p_x = data.count(x) / len(data)
            if p_x > 0:
                entropy += -p_x * np.log2(p_x)
        
        return entropy
    
    def guess_file_type(self, data):
        """Guess file type from binary data"""
        # Check for common file signatures
        signatures = {
            b'\\x50\\x4b\\x03\\x04': 'ZIP archive',
            b'\\x1f\\x8b\\x08': 'GZIP compressed',
            b'\\x42\\x5a\\x68': 'BZIP2 compressed',
            b'\\x37\\x7a\\xbc\\xaf\\x27\\x1c': '7Z archive',
            b'\\x52\\x61\\x72\\x21\\x1a\\x07\\x00': 'RAR archive',
            b'\\x25\\x50\\x44\\x46': 'PDF document',
            b'\\x4d\\x5a': 'Windows executable',
            b'\\x7f\\x45\\x4c\\x46': 'ELF executable',
            b'\\xca\\xfe\\xba\\xbe': 'Java class',
            b'\\xd0\\xcf\\x11\\xe0\\xa1\\xb1\\x1a\\xe1': 'Microsoft document',
            b'\\x00\\x00\\x01\\x00': 'Windows icon',
            b'\\x89\\x50\\x4e\\x47': 'PNG image',
            b'\\xff\\xd8\\xff': 'JPEG image',
            b'\\x47\\x49\\x46\\x38': 'GIF image',
            b'\\x49\\x44\\x33': 'MP3 audio',
            b'\\x00\\x00\\x00\\x18\\x66\\x74\\x79\\x70': 'MP4 video'
        }
        
        for sig, filetype in signatures.items():
            if data.startswith(sig):
                return filetype
        
        return 'Unknown binary'
    
    def is_archive(self, file_path):
        """Check if file is an archive"""
        archive_exts = {'.zip', '.tar', '.gz', '.bz2', '.xz', '.7z', '.rar', '.tgz'}
        return file_path.suffix.lower() in archive_exts
    
    def extract_and_process(self, file_path):
        """Extract archive and process contents"""
        print(f"[+] Extracting archive: {file_path}")
        self.process_archive(file_path)
